"""Module with the functionality of extract the data from the dump
"""
import pandas as pd
from openpyxl import load_workbook


def extract_data(file=None):
    """Function to extract the data from the engine"""

    try:
        # source_wb = '/mnt/c/Users/zrvic/OneDrive/Escritorio/Openpyxl/2020.06.05_Motor_Holberton.xlsx'
        wb_source = load_workbook(file, data_only=True)
        ws1 = wb_source.worksheets[1]
        dump = []
        for item in ws1.iter_rows(
            min_row=ws1.min_row,
            max_row=ws1.max_row,
            min_col=ws1["R1"].col_idx,
            max_col=ws1["AZ1"].col_idx,
            values_only=True,
        ):
            dump.append(item)
        return dump
    except Exception as e:
        print("{}".format(e))
        return False
